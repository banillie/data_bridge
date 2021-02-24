# import csv
import datetime

# import difflib
# import os
# import pickle
# import re
# import sys
# import typing
# from collections import Counter
import math
from typing import List, Dict, Union, Optional, Tuple
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from datetime import timedelta, date
#
# from dateutil import parser
import numpy as np
from datamaps.api import project_data_from_master
import platform
from pathlib import Path

# from dateutil.parser import ParserError
from docx import Document, table
# from docx.enum.section import WD_SECTION_START, WD_ORIENTATION
# from docx.enum.text import WD_ALIGN_PARAGRAPH
# from docx.oxml import parse_xml
# from docx.oxml.ns import nsdecls
# from docx.shared import Pt, Cm, RGBColor, Inches
# from matplotlib import cm
# from matplotlib.patches import Wedge, Rectangle, Circle
from openpyxl import load_workbook, Workbook

# from openpyxl.chart import BubbleChart, Reference, Series
from openpyxl.formatting import Rule
from openpyxl.styles import Font, PatternFill, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.workbook import workbook
from textwrap import wrap
import logging
from pdf2image import convert_from_path
from analysis_engine.data import (
    convert_bc_stage_text,
    plus_minus_days,
    concatenate_dates,
    convert_rag_text,
    rag_txt_list,
    black_text, fill_colour_list, get_group, COLOUR_DICT, make_file_friendly, wd_heading, key_contacts, dca_table,
    dca_narratives, open_word_doc,
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s: %(levelname)s - %(message)s",
    datefmt="%d-%b-%y %H:%M:%S",
)
logger = logging.getLogger(__name__)


class ProjectNameError(Exception):
    pass


def _platform_docs_dir() -> Path:
    #  Cross plaform file path handling
    if platform.system() == "Linux":
        return Path.home() / "Documents" / "data_bridge"
    if platform.system() == "Darwin":
        return Path.home() / "Documents" / "data_bridge"
    else:
        return Path.home() / "Documents" / "data_bridge"


root_path = _platform_docs_dir()


def get_master_data() -> List[
    Dict[str, Union[str, int, datetime.date, float]]
]:  # how specify a list of dictionaries?
    """Returns a list of dictionaries each containing quarter data"""
    master_data_list = [
        project_data_from_master(
            root_path / "core_data/cdg_master_3_2020.xlsx", 3, 2020
        ),
    ]
    return master_data_list


def get_project_information() -> Dict[str, Union[str, int]]:
    """Returns dictionary containing all project meta data"""
    return project_data_from_master(
        root_path / "core_data/cdg_project_info.xlsx", 2, 2020
    )


def place_data_into_new_master_format(master_data: Dict):  # throw away
    wb = load_workbook(root_path / "core_data/CDG_portfolio_report.xlsx")
    ws = wb.active

    for i, p in enumerate(master_data.projects):
        ws.cell(row=3, column=i + 5).value = p
        for row_num in range(2, ws.max_row + 1):
            key = ws.cell(row=row_num, column=2).value
            try:
                ws.cell(row=row_num, column=i + 5).value = master_data.data[p][key]
            except KeyError:
                pass

    return wb


BASELINE_TYPES = {
    "Re-baseline this quarter": "quarter",
}
CDG_GROUP_DICT = {"Corporate Finance": "CF", "Group Finance": "GF"}
BC_STAGE_DICT = {
    "Strategic Outline Case": "SOBC",
    "SOBC": "SOBC",
    "pre-Strategic Outline Case": "pre-SOBC",
    "pre-SOBC": "pre-SOBC",
    "Outline Business Case": "OBC",
    "OBC": "OBC",
    "Full Business Case": "FBC",
    "FBC": "FBC",
    # older returns that require cleaning
    "Pre - SOBC": "pre-SOBC",
    "Pre Strategic Outline Business Case": "pre_SOBC",
    None: None,
    "Other": "Other",
    "Other ": "Other",
    "To be confirmed": None,
    "To be confirmed ": None,
}
DCG_DATE = datetime.date(
    2021, 2, 22
)  # ipdc date. Python date format is Year, Month, day


class Master:
    def __init__(
        self,
        master_data: List[Dict[str, Union[str, int, datetime.date, float]]],
        project_information: Dict[str, Union[str, int]],
    ) -> None:
        self.master_data = master_data
        self.project_information = project_information
        self.current_quarter = self.master_data[0].quarter
        self.current_projects = self.master_data[0].projects
        self.abbreviations = {}
        self.full_names = {}
        self.bl_info = {}
        self.bl_index = {}
        self.dft_groups = {}
        self.project_group = {}
        self.project_stage = {}
        self.quarter_list = []
        self.get_quarter_list()
        self.get_baseline_data()
        self.check_project_information()
        self.get_project_abbreviations()
        self.check_baselines()
        self.get_project_groups()

    def get_project_abbreviations(self) -> None:
        """gets the abbreviations for all current projects.
        held in the project info document"""
        abb_dict = {}
        fn_dict = {}
        error_case = []
        for p in self.project_information.projects:
            abb = self.project_information[p]["Abbreviations"]
            abb_dict[p] = {"abb": abb, "full name": p}
            fn_dict[abb] = p
            if abb is None:
                error_case.append(p)

        if error_case:
            for p in error_case:
                logger.critical("No abbreviation provided for " + p + ".")
            raise ProjectNameError(
                "Abbreviations must be provided for all projects in project_info. Program stopping. Please amend"
            )

        self.abbreviations = abb_dict
        self.full_names = fn_dict

    def get_baseline_data(self) -> None:
        """
        Returns the two dictionaries baseline_info and baseline_index for all projects for all
        baseline types
        """

        baseline_info = {}
        baseline_index = {}

        for b_type in list(BASELINE_TYPES.keys()):
            project_baseline_info = {}
            project_baseline_index = {}
            for name in self.current_projects:
                bc_list = []
                lower_list = []
                for i, master in reversed(list(enumerate(self.master_data))):
                    if name in master.projects:
                        try:
                            approved_bc = master.data[name][b_type]
                            quarter = str(master.quarter)
                        # exception handling in here in case data keys across masters are not consistent.
                        except KeyError:
                            print(
                                str(b_type)
                                + " keys not present in "
                                + str(master.quarter)
                            )
                        if approved_bc == "YES":
                            bc_list.append(approved_bc)
                            lower_list.append((approved_bc, quarter, i))
                    else:
                        pass
                for i in reversed(range(2)):
                    try:
                    # if name in self.master_data[i].projects:
                        approved_bc = self.master_data[i][name][b_type]
                        quarter = str(self.master_data[i].quarter)
                        lower_list.append((approved_bc, quarter, i))
                    # TODO tidy this
                    except IndexError:
                    # else:
                    #     quarter = str(self.master_data[i].quarter)
                        lower_list.append((None, "LAST", None))

                index_list = []
                for x in lower_list:
                    index_list.append(x[2])

                project_baseline_info[name] = list(reversed(lower_list))
                project_baseline_index[name] = list(reversed(index_list))

            baseline_info[BASELINE_TYPES[b_type]] = project_baseline_info
            baseline_index[BASELINE_TYPES[b_type]] = project_baseline_index

        self.bl_info = baseline_info
        self.bl_index = baseline_index

    def check_project_information(self) -> None:
        """Checks that project names in master are present/the same as in project info.
        Stops the programme if not"""
        error_cases = []
        for p in self.current_projects:
            if p not in self.project_information.projects:
                error_cases.append(p)

        if error_cases:
            for p in error_cases:
                logger.critical(p + " has not been found in the project_info document.")
            raise ProjectNameError(
                "Project names in "
                + str(self.master_data[0].quarter)
                + " master and project_info must match. Program stopping. Please amend."
            )
        else:
            logger.info("The latest master and project information match")

    def check_baselines(self) -> None:
        """checks that projects have the correct baseline information. stops the
        programme if baselines are missing"""
        # work through best way to stop the programme.
        for v in BASELINE_TYPES.values():
            for p in self.current_projects:
                baselines = self.bl_index[v][p]
                if len(baselines) <= 2:
                    print(
                        p
                        + " does not have a baseline point for "
                        + v
                        + " this could cause the programme to "
                        "crash. Therefore the programme is stopping. "
                        "Please amend the data for " + p + " so that "
                        " it has at least one baseline point for " + v
                    )
            else:
                continue
            break

    def get_project_groups(self) -> None:
        """gets the groups that projects are part of e.g. business case
        stage or dft group"""

        raw_dict = {}
        raw_list = []
        group_list = []
        stage_list = []
        for i, master in enumerate(self.master_data):
            lower_dict = {}
            for p in master.projects:
                try:
                    dft_group = CDG_GROUP_DICT[
                        master[p]["CDG Group"]
                    ]  # different groups cleaned here
                    stage = BC_STAGE_DICT[master[p]["CDG approval point"]]
                    raw_list.append(("group", dft_group))
                    raw_list.append(("stage", stage))
                    lower_dict[p] = dict(raw_list)
                    group_list.append(dft_group)
                    stage_list.append(stage)
                except KeyError:
                    print(
                        str(master.quarter)
                        + ": "
                        + str(p)
                        + " has reported an incorrect DfT Group value. Amend"
                    )
            raw_dict[str(master.quarter)] = lower_dict

        group_list = list(set(group_list))
        stage_list = list(set(stage_list))

        group_dict = {}
        for i, quarter in enumerate(raw_dict.keys()):
            lower_g_dict = {}
            for group_type in group_list:
                g_list = []
                for p in raw_dict[quarter].keys():
                    p_group = raw_dict[quarter][p]["group"]
                    if p_group == group_type:
                        g_list.append(p)
                # messaging to clean up group data.
                # TODO wrap into system messaging
                if group_type is None or group_type == "DfT":
                    if g_list:
                        for x in g_list:
                            print(
                                str(quarter)
                                + " "
                                + str(x)
                                + " DfT Group data needs cleaning. Currently "
                                + str(group_type)
                            )
                lower_g_dict[group_type] = g_list

            group_dict[quarter] = lower_g_dict

        stage_dict = {}
        for quarter in raw_dict.keys():
            lower_s_dict = {}
            for stage_type in stage_list:
                s_list = []
                for p in raw_dict[quarter].keys():
                    p_stage = raw_dict[quarter][p]["stage"]
                    if p_stage == stage_type:
                        s_list.append(p)
                # messaging to clean up group data.
                # TODO wrap into system messaging
                if stage_type is None:
                    if s_list:
                        for x in s_list:
                            print(
                                str(quarter)
                                + " "
                                + str(x)
                                + " IPDC stage data needs cleaning. Currently "
                                + str(stage_type)
                            )
                lower_s_dict[stage_type] = s_list
            stage_dict[quarter] = lower_s_dict

        self.dft_groups = group_dict
        self.project_stage = stage_dict

    def get_quarter_list(self) -> None:
        output_list = []
        for master in self.master_data:
            output_list.append(str(master.quarter))
        self.quarter_list = output_list


def overall_dashboard(master: Master, wb: Workbook) -> Workbook:
    wb = load_workbook(wb)
    ws = wb.worksheets[0]

    for row_num in range(2, ws.max_row + 1):
        project_name = ws.cell(row=row_num, column=2).value
        if project_name in master.current_projects:
            """BC Stage"""
            bc_stage = master.master_data[0].data[project_name]["CDG approval point"]
            # ws.cell(row=row_num, column=4).value = convert_bc_stage_text(bc_stage)
            ws.cell(row=row_num, column=3).value = convert_bc_stage_text(bc_stage)
            try:
                bc_stage_lst_qrt = master.master_data[1].data[project_name][
                    "CDG approval point"
                ]
                if bc_stage != bc_stage_lst_qrt:
                    # ws.cell(row=row_num, column=4).font = Font(
                    #     name="Arial", size=10, color="00fc2525"
                    # )
                    ws.cell(row=row_num, column=3).font = Font(
                        name="Arial", size=10, color="00fc2525"
                    )
            except (KeyError, IndexError):
                pass

            """planning stage"""
            plan_stage = master.master_data[0].data[project_name]["Project stage"]
            # ws.cell(row=row_num, column=5).value = plan_stage
            ws.cell(row=row_num, column=4).value = plan_stage
            try:
                plan_stage_lst_qrt = master.master_data[1].data[project_name][
                    "Project stage"
                ]
                if plan_stage != plan_stage_lst_qrt:
                    # ws.cell(row=row_num, column=5).font = Font(
                    #     name="Arial", size=10, color="00fc2525"
                    # )
                    ws.cell(row=row_num, column=4).font = Font(
                        name="Arial", size=10, color="00fc2525"
                    )
            except (KeyError, IndexError):
                pass

            """Total WLC"""
            wlc_now = master.master_data[0].data[project_name]["Total Forecast"]
            # ws.cell(row=row_num, column=6).value = wlc_now
            ws.cell(row=row_num, column=5).value = wlc_now
            """WLC variance against lst quarter"""
            try:
                wlc_lst_quarter = master.master_data[1].data[project_name][
                    "Total Forecast"
                ]
                diff_lst_qrt = wlc_now - wlc_lst_quarter
                if float(diff_lst_qrt) > 0.49 or float(diff_lst_qrt) < -0.49:
                    # ws.cell(row=row_num, column=7).value = diff_lst_qrt
                    ws.cell(row=row_num, column=6).value = diff_lst_qrt
                else:
                    # ws.cell(row=row_num, column=7).value = "-"
                    ws.cell(row=row_num, column=6).value = "-"

                try:
                    percentage_change = ((wlc_now - wlc_lst_quarter) / wlc_now) * 100
                    if percentage_change > 5 or percentage_change < -5:
                        # ws.cell(row=row_num, column=7).font = Font(
                        #     name="Arial", size=10, color="00fc2525"
                        # )
                        ws.cell(row=row_num, column=6).font = Font(
                            name="Arial", size=10, color="00fc2525"
                        )
                except ZeroDivisionError:
                    pass

            except (KeyError, IndexError):
                ws.cell(row=row_num, column=6).value = "-"

            """WLC variance against baseline quarter"""
            bl = master.bl_index["quarter"][project_name][2]
            wlc_baseline = master.master_data[bl].data[project_name]["Total Forecast"]
            try:
                diff_bl = wlc_now - wlc_baseline
                if float(diff_bl) > 0.49 or float(diff_bl) < -0.49:
                    # ws.cell(row=row_num, column=8).value = diff_bl
                    ws.cell(row=row_num, column=7).value = diff_bl
                else:
                    # ws.cell(row=row_num, column=8).value = "-"
                    ws.cell(row=row_num, column=7).value = "-"
            except TypeError:  # exception is here as some projects e.g. Hs2 phase 2b have (real) written into historical totals
                pass

            try:
                percentage_change = ((wlc_now - wlc_baseline) / wlc_now) * 100
                if percentage_change > 5 or percentage_change < -5:
                    # ws.cell(row=row_num, column=8).font = Font(
                    #     name="Arial", size=10, color="00fc2525"
                    # )
                    ws.cell(row=row_num, column=7).font = Font(
                        name="Arial", size=10, color="00fc2525"
                    )

            except (
                ZeroDivisionError,
                TypeError,
            ):  # zerodivision error obvious, type error handling as above
                pass

            """vfm category now"""
            vfm_cat = master.master_data[0].data[project_name]["VfM Category single entry"]
            # if (
            #     master.master_data[0].data[project_name]["VfM Category single entry"]
            #     is None
            # ):
            #     vfm_cat = (
            #         str(
            #             master.master_data[0].data[project_name][
            #                 "VfM Category lower range"
            #             ]
            #         )
            #         + " - "
            #         + str(
            #             master.master_data[0].data[project_name][
            #                 "VfM Category upper range"
            #             ]
            #         )
            #     )
            #     # ws.cell(row=row_num, column=10).value = vfm_cat
            #     ws.cell(row=row_num, column=8).value = vfm_cat
            #
            # else:
            #     vfm_cat = master.master_data[0].data[project_name][
            #         "VfM Category single entry"
            #     ]
            #     # ws.cell(row=row_num, column=10).value = vfm_cat
            ws.cell(row=row_num, column=8).value = vfm_cat

            """vfm category baseline"""
            bl_i = master.bl_index["quarter"][project_name][2]
            vfm_cat_baseline = master.master_data[bl_i].data[project_name][
                "VfM Category single entry"
            ]
            # try:
            #     if (
            #         master.master_data[bl_i].data[project_name][
            #             "VfM Category single entry"
            #         ]
            #         is None
            #     ):
            #         vfm_cat_baseline = (
            #             str(
            #                 master.master_data[bl_i].data[project_name][
            #                     "VfM Category lower range"
            #                 ]
            #             )
            #             + " - "
            #             + str(
            #                 master.master_data[bl_i].data[project_name][
            #                     "VfM Category upper range"
            #                 ]
            #             )
            #         )
            #         # ws.cell(row=row_num, column=11).value = vfm_cat_baseline
            #     else:
            #         vfm_cat_baseline = master.master_data[bl_i].data[project_name][
            #             "VfM Category single entry"
            #         ]
            #         # ws.cell(row=row_num, column=11).value = vfm_cat_baseline

            # except KeyError:
            #     try:
            #         vfm_cat_baseline = master.master_data[bl_i].data[project_name][
            #             "VfM Category single entry"
            #         ]
            #         # ws.cell(row=row_num, column=11).value = vfm_cat_baseline
            #     except KeyError:
            #         vfm_cat_baseline = master.master_data[bl_i].data[project_name][
            #             "VfM Category"
            #         ]
            #         # ws.cell(row=row_num, column=11).value = vfm_cat_baseline

            if vfm_cat != vfm_cat_baseline:
                if vfm_cat_baseline is None:
                    pass
                else:
                    ws.cell(row=row_num, column=8).font = Font(
                        name="Arial", size=8, color="00fc2525"
                    )

            current = master.master_data[0].data[project_name]["Project End Date"]
            try:
                last_quarter = master.master_data[1].data[project_name][
                    "Full Operations"
                ]
            except IndexError:
                pass
            bl = master.master_data[bl_i].data[project_name]["Project End Date"]
            #
            # abb = master.abbreviations[project_name]["abb"]
            # current = get_milestone_date(
            #     abb, milestones.milestone_dict, "current", " Full Operations"
            # )
            # last_quarter = get_milestone_date(
            #     abb, milestones.milestone_dict, "last", " Full Operations"
            # )
            # bl = get_milestone_date(
            #     abb, milestones.milestone_dict, "bl_one", " Full Operations"
            # )
            ws.cell(row=row_num, column=9).value = current
            if current is not None and current < DCG_DATE:
                ws.cell(row=row_num, column=9).value = "Completed"
            try:
                last_change = (current - last_quarter).days
                if last_change == 0:
                    ws.cell(row=row_num, column=10).value = '-'
                else:
                    ws.cell(row=row_num, column=10).value = plus_minus_days(last_change)
                if last_change is not None and last_change > 46:
                    ws.cell(row=row_num, column=10).font = Font(
                        name="Arial", size=10, color="00fc2525"
                    )
            except (TypeError, UnboundLocalError):
                pass
            try:
                bl_change = (current - bl).days
                if bl_change == 0:
                    ws.cell(row=row_num, column=11).value = '-'
                else:
                    ws.cell(row=row_num, column=11).value = plus_minus_days(bl_change)
                if bl_change is not None and bl_change > 85:
                    ws.cell(row=row_num, column=11).font = Font(
                        name="Arial", size=10, color="00fc2525"
                    )
            except TypeError:
                pass

            # last at/next at cdg information  removed
            try:
                ws.cell(row=row_num, column=12).value = concatenate_dates(
                    master.master_data[0].data[project_name]["Last date at CDG"],
                    DCG_DATE,
                )
                ws.cell(row=row_num, column=13).value = concatenate_dates(
                    master.master_data[0].data[project_name]["Next date at CDG"],
                    DCG_DATE,
                )
            except (KeyError, TypeError):
                print(
                    project_name
                    + " last at / next at ipdc data could not be calculated. Check data."
                )

            # """IPA DCA rating"""
            # ipa_dca = convert_rag_text(
            #     master.master_data[0].data[project_name]["GMPP - IPA DCA"]
            # )
            # ws.cell(row=row_num, column=15).value = ipa_dca
            # if ipa_dca == "None":
            #     ws.cell(row=row_num, column=15).value = ""

            """DCA rating - this quarter"""
            ws.cell(row=row_num, column=17).value = convert_rag_text(
                master.master_data[0].data[project_name]["Departmental DCA"]
            )
            """DCA rating - last qrt"""
            try:
                ws.cell(row=row_num, column=19).value = convert_rag_text(
                    master.master_data[1].data[project_name]["Departmental DCA"]
                )
            except (KeyError, IndexError):
                ws.cell(row=row_num, column=19).value = ""
            """DCA rating - 2 qrts ago"""
            try:
                ws.cell(row=row_num, column=20).value = convert_rag_text(
                    master.master_data[2].data[project_name]["Departmental DCA"]
                )
            except (KeyError, IndexError):
                ws.cell(row=row_num, column=20).value = ""
            """DCA rating - 3 qrts ago"""
            try:
                ws.cell(row=row_num, column=21).value = convert_rag_text(
                    master.master_data[3].data[project_name]["Departmental DCA"]
                )
            except (KeyError, IndexError):
                ws.cell(row=row_num, column=21).value = ""
            """DCA rating - baseline"""
            bl_i = master.bl_index["quarter"][project_name][2]
            ws.cell(row=row_num, column=23).value = convert_rag_text(
                master.master_data[bl_i].data[project_name]["Departmental DCA"]
            )

        """list of columns with conditional formatting"""
        list_columns = ["o", "q", "s", "t", "u", "w"]

        """same loop but the text is black. In addition these two loops go through the list_columns list above"""
        for column in list_columns:
            for i, dca in enumerate(rag_txt_list):
                text = black_text
                fill = fill_colour_list[i]
                dxf = DifferentialStyle(font=text, fill=fill)
                rule = Rule(
                    type="containsText", operator="containsText", text=dca, dxf=dxf
                )
                for_rule_formula = 'NOT(ISERROR(SEARCH("' + dca + '",' + column + "5)))"
                rule.formula = [for_rule_formula]
                ws.conditional_formatting.add(column + "5:" + column + "60", rule)

        for row_num in range(2, ws.max_row + 1):
            for col_num in range(5, ws.max_column + 1):
                if ws.cell(row=row_num, column=col_num).value == 0:
                    ws.cell(row=row_num, column=col_num).value = "-"

    return wb


class DandelionData:
    def __init__(self, master: Master, **kwargs):
        self.master = master
        self.kwargs = kwargs
        self.baseline_type = "ipdc_costs"
        self.group = []
        self.iter_list = []
        self.d_data = {}
        self.d_list = []
        self.get_data()

    def get_data(self):
        self.group = self.master.master_data[0].projects
        input_g_list = ["CF", "GF"]  # first outer circle
        # input_g_list = ["FBC", "OBC", "SOBC", "pre-SOBC"]  # first outer circle
        # cal group angle
        if len(input_g_list) == 2:
            g_ang_list = [0, 225, 315]
        # g_ang = 270/len(input_g_list)  # group angle
        # g_ang_list = []
        # for i in range(6):
        #     g_ang_list.append(g_ang * i)
        # del g_ang_list[4]

        dft_g_list = []
        dft_g_dict = {}
        dft_l_group_dict = {}
        p_total = 0  # portfolio total
        for i, g in enumerate(input_g_list):
            dft_l_group = self.master.dft_groups["Q3 20/21"][g]
            g_total = 0
            dft_l_group_list = []
            for p in dft_l_group:
                p_data = self.master.master_data[0].data[p]
                b_size = p_data["Total Forecast"]
                if b_size is None:
                    b_size = 25
                rag = p_data["Departmental DCA"]
                colour = COLOUR_DICT[convert_rag_text(rag)]
                g_total += b_size
                dft_l_group_list.append((math.sqrt(b_size), colour, self.master.abbreviations[p]['abb']))
            # group data
            x_axis = 0 + 120 * math.cos(math.radians(g_ang_list[i + 1]))
            y_axis = 0 + 100 * math.sin(math.radians(g_ang_list[i + 1]))
            # list is tuple axis point, bubble size, colour, line style, line color, text position
            dft_g_list.append(((x_axis, y_axis),
                               math.sqrt(g_total),
                               '#a0c1d5',
                               g,
                               'dashed',
                               'grey',
                               ('center', 'center')))
            dft_g_dict[g] = [(x_axis, y_axis), math.sqrt(g_total)/3]  # used for placement of circles
            # project data
            dft_l_group_dict[g] = list(reversed(sorted(dft_l_group_list)))
            #portfolio data
            p_total += g_total
        dft_g_list.append(((0, 0), math.sqrt(p_total), '#a0c1d5', "CDG Portfolio", "dashed", "grey", ('center', 'center')))

        for g in dft_l_group_dict.keys():
            lg = dft_l_group_dict[g]  # local group
            ang = 360/len(lg)
            ang_list = []
            for i in range(len(lg)+1):
                ang_list.append(ang*i)
            for i, p in enumerate(lg):
                a = dft_g_dict[g][0][0]
                b = dft_g_dict[g][0][1]
                if len(lg) <= 8:
                    x_axis = a + (dft_g_dict[g][1] + 20) * math.cos(math.radians(ang_list[i+1]))
                    y_axis = b + (dft_g_dict[g][1] + 20) * math.sin(math.radians(ang_list[i+1]))
                else:
                    x_axis = a + (dft_g_dict[g][1] + 40) * math.cos(math.radians(ang_list[i + 1]))
                    y_axis = b + (dft_g_dict[g][1] + 40) * math.sin(math.radians(ang_list[i + 1]))
                b_size = p[0]
                colour = p[1]
                name = p[2]
                if 280 >= ang_list[i+1] >= 80:
                    text_angle = ('right', 'bottom')
                if 100 >= ang_list[i+1] or ang_list[i+1] >= 260:
                    text_angle = ('left', 'bottom')
                if 279 >= ang_list[i+1] >= 261:
                    text_angle = ('center', 'top')
                if 99 >= ang_list[i+1] >= 81:
                    text_angle = ('center', 'bottom')
                dft_g_list.append(((x_axis, y_axis), b_size, colour, name, "solid", colour, text_angle))

        self.d_list = dft_g_list


def make_a_dandelion_auto_cdg(dlion_data: DandelionData):
    fig, ax = plt.subplots()
    ax.set_facecolor('#a0c1d5')
    for c in range(len(dlion_data.d_list)):
        circle = plt.Circle(dlion_data.d_list[c][0],
                            radius=dlion_data.d_list[c][1],
                            fc=dlion_data.d_list[c][2],  # face colour
                            linestyle=dlion_data.d_list[c][4],
                            ec=dlion_data.d_list[c][5])  # edge colour
        ax.add_patch(circle)
        ax.annotate(dlion_data.d_list[c][3],
                    xy=dlion_data.d_list[c][0],
                    fontsize=6,
                    horizontalalignment=dlion_data.d_list[c][6][0],
                    verticalalignment=dlion_data.d_list[c][6][1])
        # plt.gca().add_patch(circle)

    plt.axis('scaled')
    plt.axis('off')
    plt.show()

    return plt


def convert_pdf_to_png():
    pages = convert_from_path(root_path / "output/dandelion.pdf", 500)
    for page in pages:
        page.save(root_path / "output/dandelion.jpeg", 'JPEG')


def compile_p_report_cdg(
        doc: Document,
        project_info: Dict[str, Union[str, int, date, float]],
        master: Master,
        project_name: str,
) -> Document:
    wd_heading(doc, project_info, project_name)
    key_contacts(doc, master, project_name)
    dca_table(doc, master, project_name)
    dca_narratives(doc, master, project_name)
    # costs = CostData(master, group=[project_name], baseline=["standard"])
    # benefits = BenefitsData(master, project_name)
    # milestones = MilestoneData(master, group=[project_name], baseline=["standard"])
    # project_report_meta_data(doc, costs, milestones, benefits, project_name)
    # change_word_doc_landscape(doc)
    # cost_profile = cost_profile_graph(costs, show="No")
    # put_matplotlib_fig_into_word(doc, cost_profile, transparent=False, size=8)
    # total_profile = total_costs_benefits_bar_chart(costs, benefits, show="No")
    # put_matplotlib_fig_into_word(doc, total_profile, transparent=False, size=8)
    # #  handling of no milestones within filtered period.
    # ab = master.abbreviations[project_name]["abb"]
    # try:
    #     # milestones.get_milestones()
    #     # milestones.get_chart_info()
    #     milestones.filter_chart_info(dates=["1/9/2020", "30/12/2022"])
    #     milestones_chart = milestone_chart(
    #         milestones,
    #         blue_line="ipdc_date",
    #         title=ab + " schedule (2021 - 22)",
    #         show="No",
    #     )
    #     put_matplotlib_fig_into_word(doc, milestones_chart, transparent=False, size=8)
    #     # print_out_project_milestones(doc, milestones, project_name)
    # except ValueError:  # extends the time period.
    #     milestones = MilestoneData(master, project_name)
    #     # milestones.get_milestones()
    #     # milestones.get_chart_info()
    #     milestones.filter_chart_info(dates=["1/9/2020", "30/12/2024"])
    #     milestones_chart = milestone_chart(
    #         milestones,
    #         blue_line="ipdc_date",
    #         title=ab + " schedule (2021 - 24)",
    #         show="No",
    #     )
    #     put_matplotlib_fig_into_word(doc, milestones_chart)
    # print_out_project_milestones(doc, milestones, project_name)
    # change_word_doc_portrait(doc)
    # project_scope_text(doc, master, project_name)
    return doc


def run_p_reports_cdg(master: Master, **kwargs) -> None:
    group = master.current_projects
    # group = get_group(master, str(master.current_quarter), kwargs)

    for p in group:
        print("Compiling summary for " + p)
        report_doc = open_word_doc(root_path / "input/summary_temp.docx")
        qrt = make_file_friendly(str(master.master_data[0].quarter))
        output = compile_p_report_cdg(report_doc, get_project_information(), master, p)
        abb = master.abbreviations[p]["abb"]
        output.save(
            root_path / "output/{}_report_{}.docx".format(abb, qrt)
        )  # add quarter here

